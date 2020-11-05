from openpyxl import load_workbook
import os


# filelist: demographics / procedure


def demoxlsx2csv(fname, output_fname):
    wb = load_workbook(filename=fname)
    sheet = wb.active
    line = ''
    for row in sheet.rows:
        _line = []
        for cell in row:
            _line.append(str(cell.value))
            _line.append(',')
        _line[-1] = '\n'
        line += ''.join(_line)
    
    with open(output_fname, 'wt') as f:
        f.writelines(line[:-1])


def proc2csv(fname, output_fname):
    wb = load_workbook(filename=fname)

    def _proc(sheet, ofname):
        line = ''
        for row in sheet.rows:
            _line = []
            for i, cell in enumerate(row):
                if i >= 3:
                    continue
                _line.append(str(cell.value))
                _line.append(',')
            _line[-1] = '\n'
            line += ''.join(_line)
        
        with open(ofname, 'wt') as f:
            f.writelines(line[:-1])
            
    _proc(wb['position change'], os.path.join(output_fname, 'position_change.csv'))
    _proc(wb['pressure reducing device'], os.path.join(output_fname, 'pressure_reducing_device.csv'))
    

if __name__=='__main__':
#    fname = '/home/mike/codes/predicting_pressure_ulcer/dataset/1026/demographics_20201105.xlsx'
#    output_fname = '/home/mike/codes/predicting_pressure_ulcer/preprocessed/demographics.csv'
#    demoxlsx2csv(fname, output_fname)  
    fname = '/home/mike/codes/predicting_pressure_ulcer/dataset/1026/procedure_20201105.xlsx'
    output_fname = '/home/mike/codes/predicting_pressure_ulcer/preprocessed'
    proc2csv(fname, output_fname)

